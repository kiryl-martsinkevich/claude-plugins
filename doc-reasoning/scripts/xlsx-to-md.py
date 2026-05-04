#!/usr/bin/env python3
"""Extract data from Excel (.xlsx) files as markdown tables.

Usage: xlsx-to-md.py <input.xlsx> <output.md>
"""

import sys
from pathlib import Path


def extract_sheets(filepath: str) -> list[tuple[str, list[list[str]]]]:
    try:
        import openpyxl
    except ImportError:
        sys.stderr.write(
            "ERROR: openpyxl not installed. Run: pip install openpyxl\n"
        )
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets = []

    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(max_row=min(ws.max_row, 2000)):
            cells = []
            for cell in row:
                val = cell.value
                if val is None:
                    cells.append("")
                else:
                    cells.append(str(val).strip())
            # Skip fully empty rows
            if any(c for c in cells):
                rows.append(cells)
        if rows:
            sheets.append((name, rows))

    wb.close()
    return sheets


def rows_to_markdown(sheet_name: str, rows: list[list[str]]) -> str:
    """Convert rows to a GitHub-flavored markdown table."""
    if not rows:
        return ""

    # Pad rows to same column count
    max_cols = max(len(r) for r in rows)
    padded = [r + [""] * (max_cols - len(r)) for r in rows]

    lines = [f"## {sheet_name}\n"]
    # Header from first row
    header = padded[0]
    lines.append("| " + " | ".join(h or " " for h in header) + " |")
    lines.append("|" + "|".join(" --- " for _ in header) + "|")
    for row in padded[1:]:
        lines.append("| " + " | ".join(c or " " for c in row) + " |")
    lines.append("")
    return "\n".join(lines)


def main():
    if len(sys.argv) != 3:
        print(f"Usage: {sys.argv[0]} <input.xlsx> <output.md>", file=sys.stderr)
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2]

    if not Path(input_path).exists():
        print(f"ERROR: File not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    sheets = extract_sheets(input_path)

    if not sheets:
        print(f"WARNING: No data extracted from {input_path}", file=sys.stderr)
        Path(output_path).write_text(f"*(No data extracted from {Path(input_path).name})*\n")
        return

    parts = []
    for name, rows in sheets:
        parts.append(rows_to_markdown(name, rows))

    output = f"# {Path(input_path).stem}\n\n" + "\n".join(parts)
    Path(output_path).write_text(output)
    print(f"{output_path} ({len(sheets)} sheet(s))")


if __name__ == "__main__":
    main()
